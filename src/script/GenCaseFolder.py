import os
from openpyxl import load_workbook
from src.common import assets_path


def read_xlsx(filename: str)-> list:
    filepath = os.path.join(assets_path, filename)
    workbook = load_workbook(filepath)
    cases = []
    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        # 遍历行
        for row in workbook[sheet_name].iter_rows():
            row_data = [cell.value for cell in row]
            cases.append(row_data)
    return cases

